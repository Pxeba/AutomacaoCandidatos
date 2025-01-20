import os
import re
import openai
import PyPDF2
from collections import defaultdict
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

# Configuração da API OpenAI
openai.api_key = '<SUA_API_OPENAI>'

filtros = """(
1- O candidato caso fornecer idade no curriculo, deve possuir entre 20 e 28 anos. Não inferir a menos que seja 100% certo q ele não está nessa faixa /n
2- Hoje é janeiro de 2025, tenha isso em mente. O candidato não pode estar em uma empresa atualmente a não ser que especifique que é freelancer ou empresa própria. Só considere como empregado atualmente se ele deixar como explícito que está em uma empresa até a data atual / presente.
3- O candidato deve ter graduação em uma faculdade de TI ou estar fazendo uma. Exemplos são: Engenharia da computação, Ciencia da computação, Engenharia de Software, etc. /n
4- O candidato não pode ter graduação iniciada ou concluída antes de 2015.) /n
5- O candidato precisa ter experiência em programação no mercado de trabalho de pelo menos 6 meses. Não considerar experiências que não são focadas em programação como infraestrutura / suporte técnico / Help Desk. Se não tiver experiencia pratica mencionada mas possuir iniciação científica(em qualquer área) tudo bem.
"""

def extrair_texto_pdf(caminho_pdf):
    """Extrai texto de um arquivo PDF."""
    try:
        with open(caminho_pdf, 'rb') as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            texto = ''
            for page_num in range(len(reader.pages)):
                texto += reader.pages[page_num].extract_text()
            return texto
    except Exception as e:
        print(f"Erro ao ler {caminho_pdf}: {e}")
        return ''


def avaliar_candidato(texto, filtro):
    """Envia o texto do PDF e o filtro para a API da OpenAI e avalia se o candidato é aprovado ou recusado."""
    try:
        chat_response = openai.chat.completions.create(
            model="gpt-4o",
            temperature=0.3,
            messages=[
                {"role": "system", "content": "Você é um assistente que ajuda a avaliar candidatos para uma vaga."},
                {"role": "user", "content": f"Hoje é outubro de 2024. Quero que comece a resposta com um OK seguido da data de hoje pq quero ter certeza que você está considerando isso ao responder, ignorar o fato q você foi treinado até 2023 porque hoje estamos em 2024. Com base nos seguintes filtros fornecidos pelo contratante: {filtro}, diga se o candidato a seguir é aprovado ou recusado e, digite antes uma descrição analisando o candidato em cada numero de filtro e bem no final após a descrição apenas o numero da justificativa caso ele tenha sido reprovado. Formato da sua resposta: <descricao> .Recusado/Aprovado. Justificativa: 1. Com base nisso avalie o candidato a seguir: <>{texto}<>"}
            ]
        )
        # Interpreta a resposta da API
        resultado = chat_response.choices[0].message.content.lower()
        return resultado
    except Exception as e:
        print(f"Erro ao avaliar candidato: {e}")
        return "erro"


def criar_ou_atualizar_planilha():
    """Cria uma planilha ou adiciona uma nova aba se já existir."""
    nome_arquivo = 'candidatos_aprovados.xlsx'
    if os.path.exists(nome_arquivo):
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.create_sheet(f"Aprovados_{len(wb.sheetnames)}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aprovados"

    # Definir as colunas
    ws.append(["Nome", "Email", "Curso de Graduação", "Período", "Idade",
              "Anos de Experiência", "Python", "Vue", "Link Indeed", "LinkedIn", "GitHub", "WhatsApp"])

    # Ajustar o tamanho de cada coluna para 100 pixels (aproximadamente 14 unidades de largura)
    for col in range(1, 13):  # Existem 12 colunas agora
        # Define a largura aproximada em unidades
        ws.column_dimensions[get_column_letter(col)].width = 32

    return wb, ws


def limpar_link(link):
    """Remove colchetes e parênteses de links formatados como Markdown."""
    if link.startswith('[') and link.endswith(')'):
        link = link.split('](')[-1].strip(')')
    return link


def formatar_numero_telefone(telefone):
    """Formata o número de telefone para o formato internacional adequado para WhatsApp."""
    # Remove parênteses, espaços e traços
    telefone = telefone.replace("(", "").replace(
        ")", "").replace("-", "").replace(" ", "")

    # Adiciona o código do país, se necessário. Aqui, assumi que todos os números são do Brasil.
    if not telefone.startswith("55"):
        telefone = "55" + telefone

    return telefone


def preencher_planilha(ws, dados):
    """Preenche uma linha da planilha com os dados do candidato."""
    ws.append([
        dados.get('nome', ''),
        dados.get('email', ''),
        dados.get('curso', ''),
        dados.get('periodo', ''),
        dados.get('idade', ''),
        dados.get('anos_exp', ''),
        dados.get('python_exp', 'Não'),
        dados.get('vue_exp', 'Não'),
        dados.get('link_indeed', ''),
        f'=HYPERLINK("{limpar_link(dados.get("linkedin", ""))}", "LinkedIn")' if dados.get(
            "linkedin") else '',
        f'=HYPERLINK("{limpar_link(dados.get("github", ""))}", "GitHub")' if dados.get(
            "github") else '',
        f'=HYPERLINK("https://wa.me/{formatar_numero_telefone(dados.get("telefone", ""))}", "WhatsApp")' if dados.get(
            "telefone") else ''

    ])


def salvar_planilha(wb):
    """Salva a planilha no arquivo"""
    wb.save('candidatos_aprovados.xlsx')


def parse_informacoes(informacoes_str):
    """Converte a string formatada em um dicionário de informações."""
    informacoes_dict = {}

    # Padrões de correspondência para cada campo
    informacoes_dict['nome'] = re.search(
        r"(?<=\*\*Nome:\*\* ).+", informacoes_str)
    informacoes_dict['email'] = re.search(
        r"(?<=\*\*Email:\*\* ).+", informacoes_str)
    informacoes_dict['curso'] = re.search(
        r"(?<=\*\*Curso de graduação:\*\* ).+", informacoes_str)
    informacoes_dict['periodo'] = re.search(
        r"(?<=\*\*Período:\*\* ).+", informacoes_str)
    informacoes_dict['idade'] = re.search(
        r"(?<=\*\*Idade:\*\* ).+", informacoes_str)
    informacoes_dict['anos_exp'] = re.search(
        r"(?<=\*\*Anos de experiência em programação no mercado de trabalho:\*\* ).+", informacoes_str)
    informacoes_dict['python_exp'] = re.search(
        r"(?<=\*\*Experiência em Python:\*\* ).+", informacoes_str)
    informacoes_dict['vue_exp'] = re.search(
        r"(?<=\*\*Experiência em Vue:\*\* ).+", informacoes_str)
    informacoes_dict['link_indeed'] = re.search(
        r"(?<=\*\*Link Indeed:\*\* ).+", informacoes_str)
    informacoes_dict['linkedin'] = re.search(
        r"(?<=\*\*LinkedIn:\*\* ).+", informacoes_str)
    informacoes_dict['github'] = re.search(
        r"(?<=\*\*GitHub:\*\* ).+", informacoes_str)
    informacoes_dict['telefone'] = re.search(
        r"(?<=\*\*Telefone:\*\* ).+", informacoes_str)

    # Converte os resultados de MatchObject para strings e limpa os dados
    for key, value in informacoes_dict.items():
        if value:
            informacoes_dict[key] = value.group().strip()
        else:
            informacoes_dict[key] = ''  # Deixa em branco se não encontrado

    return informacoes_dict


def extrair_informacoes_dos_aprovados(texto_pdf):
    """Reenvia o texto para a OpenAI para extrair informações detalhadas dos candidatos aprovados."""
    try:
        chat_response = openai.chat.completions.create(
            model="gpt-4o",
            temperature=0.3,
            messages=[
                {"role": "system", "content": "Você é um assistente que extrai informações de candidatos de currículos."},
                {"role": "user",
                    "content": f"Extraia as seguintes informações do candidato: Nome, Email, Curso de graduação (com nome), Período (se houver), Idade, Anos de experiência em programação no mercado de trabalho, Se tem experiência em Python (Sim/Não), Se tem experiência em Vue (Sim/Não), Link Indeed, LinkedIn, GitHub, Telefone. Aqui está o conteúdo do currículo: {texto_pdf}"}
            ]
        )
        resultado = chat_response.choices[0].message.content
        # Converte a resposta formatada em dicionário
        return parse_informacoes(resultado)
    except Exception as e:
        print(f"Erro ao extrair informações: {e}")
        return {}


def processar_pdfs(caminho_pasta, pasta_aprovados):
    """Processa todos os PDFs da pasta e copia os aprovados para a pasta de aprovados, além de gerar a planilha."""
    aprovados = []
    recusados = []
    justificativas_reprovacao = defaultdict(int)

    # Verifica se a pasta de aprovados existe, se não, cria.
    if not os.path.exists(pasta_aprovados):
        os.makedirs(pasta_aprovados)

    # Cria ou atualiza a planilha
    wb, ws = criar_ou_atualizar_planilha()

    for arquivo in os.listdir(caminho_pasta):
        if arquivo.endswith('.pdf'):
            caminho_pdf = os.path.join(caminho_pasta, arquivo)
            texto_pdf = extrair_texto_pdf(caminho_pdf)
            nome_candidato = os.path.splitext(arquivo)[0]

            resultado = avaliar_candidato(texto_pdf, filtros)
            if "aprovado" in resultado:
                print(f"{nome_candidato} foi APROVADO.")
                # Extrai informações detalhadas dos aprovados
                informacoes = extrair_informacoes_dos_aprovados(texto_pdf)
                preencher_planilha(ws, informacoes)
                shutil.copy(caminho_pdf, os.path.join(
                    pasta_aprovados, arquivo))
            elif "recusado" in resultado:
                justificativa = resultado.split("justificativa:")[
                    1] if "justificativa:" in resultado else "Sem justificativa específica"
                recusados.append(f"{nome_candidato} - {justificativa}")
                justificativas_reprovacao[justificativa.strip()] += 1
                print(
                    f"{nome_candidato} foi RECUSADO. Justificativa: {justificativa}")
            else:
                print(
                    f"{nome_candidato} não pôde ser avaliado corretamente. Resultado: {resultado}")

    # Salva a planilha com os aprovados
    salvar_planilha(wb)

    # Exibe o total de aprovados e recusados
    print(f"\nTotal de aprovados: {len(aprovados)}")
    print(f"Total de recusados: {len(recusados)}")

    # Exibe a quantidade de reprovação por justificativa
    print("\nResumo das justificativas de recusa:")
    for justificativa, quantidade in justificativas_reprovacao.items():
        print(
            f"Justificativa: {justificativa} - {quantidade} candidato(s) recusado(s)")


if __name__ == "__main__":
    caminho_pasta = input("Digite o caminho da pasta com os PDFs: ")
    pasta_aprovados = input(
        "Digite o nome da pasta para salvar os PDFs dos aprovados: ")

    processar_pdfs(caminho_pasta, pasta_aprovados)